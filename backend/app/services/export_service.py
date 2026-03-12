from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas import TimetableResponse

DAY_NAMES = {
    1: "Понедельник",
    2: "Вторник",
    3: "Среда",
    4: "Четверг",
    5: "Пятница",
}


def timetable_to_xlsx(tt: TimetableResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = tt.entity_name[:31] or "Timetable"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"{tt.entity_type.upper()}: {tt.entity_name}"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Урок"
    ws["A3"].font = header_font
    ws["A3"].fill = header_fill
    ws["A3"].border = border
    ws["A3"].alignment = Alignment(horizontal="center")

    for day in range(1, 6):
        col = chr(65 + day)
        cell = ws[f"{col}3"]
        cell.value = DAY_NAMES[day]
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col].width = 28

    ws.column_dimensions["A"].width = 10

    for row_idx, row in enumerate(tt.rows, start=4):
        ws[f"A{row_idx}"] = row.period
        ws[f"A{row_idx}"].font = Font(bold=True)
        ws[f"A{row_idx}"].border = border
        ws[f"A{row_idx}"].alignment = Alignment(horizontal="center")

        for day in range(1, 6):
            col = chr(65 + day)
            cell = ws[f"{col}{row_idx}"]
            entries = row.days.get(day, [])
            if entries:
                lines: list[str] = []
                for entry in entries:
                    parts = [entry.subject]
                    if entry.teacher:
                        parts.append(entry.teacher)
                    if entry.room:
                        parts.append(f"каб. {entry.room}")
                    if entry.group:
                        parts.append(f"({entry.group})")
                    lines.append(" | ".join(parts))
                cell.value = "\n".join(lines)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def timetable_to_pdf_html(tt: TimetableResponse) -> str:
    rows_html = ""
    for row in tt.rows:
        cells_html = f"<td class='period'>{row.period}<br><small>{row.time}</small></td>"
        for day in range(1, 6):
            entries = row.days.get(day, [])
            if entries:
                inner = "<br>".join(
                    f"<b>{entry.subject}</b>"
                    + (f"<br>{entry.teacher}" if entry.teacher else "")
                    + (f"<br><i>каб. {entry.room}</i>" if entry.room else "")
                    + (f"<br>({entry.group})" if entry.group else "")
                    for entry in entries
                )
            else:
                inner = ""
            cells_html += f"<td>{inner}</td>"
        rows_html += f"<tr>{cells_html}</tr>"

    day_headers = "".join(f"<th>{DAY_NAMES[day]}</th>" for day in range(1, 6))

    return f"""<!DOCTYPE html>
<html>
<head>
  <meta charset='utf-8'>
  <style>
    body {{ font-family: Arial, sans-serif; font-size: 11px; }}
    h1 {{ font-size: 16px; text-align: center; margin: 0 0 12px 0; }}
    table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
    th, td {{ border: 1px solid #333; padding: 4px; vertical-align: top; }}
    th {{ background: #CCE5FF; text-align: center; }}
    td.period {{ text-align: center; font-weight: bold; width: 70px; }}
    small {{ color: #666; }}
  </style>
</head>
<body>
  <h1>{tt.entity_name}</h1>
  <table>
    <tr><th>Урок</th>{day_headers}</tr>
    {rows_html}
  </table>
</body>
</html>"""
